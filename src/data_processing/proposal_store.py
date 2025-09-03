"""Utilities for storing proposals and execution results in the governance workbook."""
from __future__ import annotations

import pathlib
from typing import Dict, Any, TYPE_CHECKING
import json
import warnings

from utils.helpers import utc_now_iso

ROOT = pathlib.Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data"
XLSX_PATH = DATA_DIR / "input" / "PKD Governance Data.xlsx"


if TYPE_CHECKING:
    from openpyxl import Workbook  # type: ignore


def ensure_workbook() -> "Workbook":
    """Create the governance workbook with required sheets if needed.

    Returns the loaded/created workbook.  If ``openpyxl`` is unavailable,
    an informative :class:`ImportError` is raised after emitting a warning.
    When a new workbook is created, the default "Sheet" is removed and the
    workbook is saved containing exactly the sheets
    "Referenda", "Proposals", "Context", and "ExecutionResults".
    """

    try:
        from openpyxl import load_workbook, Workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - exercised in tests via monkeypatch
        warnings.warn(
            "openpyxl is required for storing governance data; install it to enable persistence",
            stacklevel=2,
        )
        raise exc

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(XLSX_PATH) if XLSX_PATH.exists() else Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    required = ["Referenda", "Proposals", "Context", "ExecutionResults"]

    for name in required:
        if name not in wb.sheetnames:
            wb.create_sheet(name)

    # Ensure only required sheets remain and are ordered as specified
    for sheet in list(wb.sheetnames):
        if sheet not in required:
            wb.remove(wb[sheet])
    wb._sheets = [wb[name] for name in required]

    wb.save(XLSX_PATH)
    return wb


def _append_row(sheet: str, row: Dict[str, Any]) -> None:
    """Append ``row`` to ``sheet`` within the governance workbook."""

    try:
        wb = ensure_workbook()
    except ImportError:
        raise
    except Exception:
        return
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)

    header = [cell.value for cell in ws[1]] if ws.max_row else []
    if not header or all(h is None for h in header):
        ws.delete_rows(1)
        header = list(row.keys())
        ws.append(header)
    else:
        for key in row.keys():
            if key not in header:
                header.append(key)
                ws.cell(row=1, column=len(header)).value = key
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=len(header)).value = ""

    ws.append([row.get(col, "") for col in header])
    wb.save(XLSX_PATH)


def _append_governance_entry(sheet: str, row: Dict[str, Any]) -> None:
    """Convenience wrapper to persist rows to ``PKD Governance Data.xlsx``."""

    _append_row(sheet, row)


def record_proposal(
    proposal_text: str,
    submission_id: str | None,
    *,
    stage: str | None = None,
    source: str | None = None,
    forecast_confidence: float | None = None,
    source_weight: float | None = None,
    score: float | None = None,
) -> None:
    """Record a generated proposal and optional submission identifier.

    Parameters
    ----------
    proposal_text:
        The text of the proposal.
    submission_id:
        Optional on-chain submission identifier.
    stage:
        Workflow stage of the proposal (``"draft"`` or ``"submitted"``).
        When provided this is persisted alongside the proposal text so that
        intermediate drafts or submissions can later be tracked. The argument
        is keyword-only to maintain backwards compatibility with existing calls.
    source:
        Optional origin of the proposal draft. When supplied this is stored
        so that downstream analysis can attribute drafts to their source.
    forecast_confidence:
        Optional confidence score output by the forecasting model.
    source_weight:
        Optional weight assigned to the draft's source when computing the
        final selection score.
    score:
        Optional selection score associated with the draft.
    """

    row = {
        "timestamp": utc_now_iso(),
        "proposal_text": proposal_text,
        "submission_id": submission_id or "",
    }
    if stage is not None:
        row["stage"] = stage
    if source is not None:
        row["source"] = source
    if forecast_confidence is not None:
        row["forecast_confidence"] = forecast_confidence
    if source_weight is not None:
        row["source_weight"] = source_weight
    if score is not None:
        row["score"] = score
    _append_governance_entry("Proposals", row)


def record_execution_result(
    status: str,
    block_hash: str,
    outcome: str,
    submission_id: str | None = None,
    extrinsic_hash: str | None = None,
    referendum_index: int | None = None,
) -> None:
    """Append governor execution details to the ``ExecutionResults`` sheet.

    If ``referendum_index`` is provided, the latest referendum data is
    fetched and stored in the ``Referenda`` sheet via
    :func:`referenda_updater.append_referendum`.
    """
    proposal_row: int | None = None
    if submission_id:
        wb = ensure_workbook()
        if "Proposals" in wb.sheetnames:
            ws = wb["Proposals"]
            header = [cell.value for cell in ws[1]]
            try:
                sub_col = header.index("submission_id") + 1
            except ValueError:
                sub_col = None
            try:
                stage_col = header.index("stage") + 1
            except ValueError:
                stage_col = None
            if sub_col is not None:
                for r in range(2, ws.max_row + 1):
                    if ws.cell(row=r, column=sub_col).value == submission_id:
                        if stage_col is not None:
                            if ws.cell(row=r, column=stage_col).value != "submitted":
                                continue
                        proposal_row = r
                        break

    row = {
        "timestamp": utc_now_iso(),
        "submission_id": submission_id or "",
        "status": status,
        "block_hash": block_hash,
        "outcome": outcome,
        "extrinsic_hash": extrinsic_hash or "",
    }
    if proposal_row is not None:
        row["proposal_row"] = proposal_row
    _append_governance_entry("ExecutionResults", row)

    if referendum_index is not None:
        try:
            from src.data_processing import referenda_updater

            referenda_updater.append_referendum(referendum_index)
        except Exception:
            pass


def record_context(context_dict: Dict[str, Any]) -> None:
    """Record consolidated context blob for auditing."""
    row = {
        "timestamp": utc_now_iso(),
        "context_json": json.dumps(context_dict),
    }
    _append_row("Context", row)


# Reading helpers -----------------------------------------------------------


def load_proposals():
    """Read the ``Proposals`` sheet as a DataFrame (empty if unavailable)."""
    import pandas as pd

    if not XLSX_PATH.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(XLSX_PATH, sheet_name="Proposals")
    except Exception:
        return pd.DataFrame()


def load_execution_results():
    """Read the ``ExecutionResults`` sheet as a DataFrame (empty if unavailable)."""
    import pandas as pd

    if not XLSX_PATH.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(XLSX_PATH, sheet_name="ExecutionResults")
    except Exception:
        return pd.DataFrame()


def load_contexts():
    """Read the ``Context`` sheet as a DataFrame (empty if unavailable)."""
    import pandas as pd

    if not XLSX_PATH.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(XLSX_PATH, sheet_name="Context")
    except Exception:
        return pd.DataFrame()


def retrieve_recent(topics: list[str], limit_per_topic: int = 3) -> list[str]:
    """Return recent proposal/context snippets mentioning any of ``topics``.

    Searches both the ``Proposals`` and ``Context`` sheets for entries
    containing any of the provided topics (case-insensitive) and returns the
    most recent snippets. Results are limited to ``limit_per_topic`` per topic
    and returned in no particular order.
    """

    if not topics:
        return []

    proposals = load_proposals()
    contexts = load_contexts()
    snippets: list[str] = []

    for topic in topics:
        if not isinstance(topic, str) or not topic:
            continue
        if not proposals.empty and "proposal_text" in proposals.columns:
            mask = proposals["proposal_text"].astype(str).str.contains(topic, case=False, na=False)
            snippets.extend(
                proposals.loc[mask]
                .sort_values("timestamp", ascending=False)
                .head(limit_per_topic)["proposal_text"].astype(str).tolist()
            )
        if not contexts.empty and "context_json" in contexts.columns:
            mask = contexts["context_json"].astype(str).str.contains(topic, case=False, na=False)
            snippets.extend(
                contexts.loc[mask]
                .sort_values("timestamp", ascending=False)
                .head(limit_per_topic)["context_json"].astype(str).tolist()
            )

    return snippets


def search_proposals(query: str, limit: int) -> list[str]:
    """Return up to ``limit`` proposal texts containing ``query``.

    Performs a case-insensitive search over the stored proposals using
    pandas' string matching and returns a list of matching snippets.
    """
    import pandas as pd

    if not query:
        return []

    df = load_proposals()
    if df.empty or "proposal_text" not in df.columns:
        return []

    mask = df["proposal_text"].astype(str).str.contains(query, case=False, na=False)
    return df.loc[mask, "proposal_text"].head(limit).tolist()
