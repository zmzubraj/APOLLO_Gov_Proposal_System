from data_processing import proposal_store
from openpyxl import load_workbook


def test_ensure_workbook_creates_required_sheets(tmp_path, monkeypatch):
    tmp_xlsx = tmp_path / "gov.xlsx"
    monkeypatch.setattr(proposal_store, "XLSX_PATH", tmp_xlsx)

    proposal_store.ensure_workbook()
    wb = load_workbook(tmp_xlsx)

    expected = {
        "Referenda",
        "DraftedProposals",
        "Proposal",
        "Context",
        "ExecutionResults",
    }
    assert set(wb.sheetnames) == expected
    assert "Sheet" not in wb.sheetnames
