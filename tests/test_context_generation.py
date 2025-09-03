import json

from src.agents.context_generator import build_context
from data_processing import proposal_store
from src.utils import validators as v
from llm import ollama_api


def _dummy_components():
    sentiment = {
        "sentiment_score": 0,
        "summary": "ok",
        "key_topics": [],
        "sentiment": "Mixed",
        "confidence": 0.0,
        "message_size_kb": 0.0,
    }
    news = {"digest": [], "risks": ""}
    chain = {
        "daily_tx_count": {},
        "daily_total_fees_DOT": {},
        "avg_tx_per_block": 0,
        "avg_fee_per_tx_DOT": 0,
        "busiest_hour_utc": "",
    }
    gov = {
        "total_referenda": 0,
        "executed_pct": 0,
        "rejected_pct": 0,
        "avg_turnout_pct": 0,
        "median_turnout_pct": 0,
        "avg_participants": 0,
        "avg_duration_days": 0,
        "monthly_counts": {},
        "top_keywords": [],
    }
    evm = {
        "daily_tx_count": {},
        "daily_total_value_ETH": {},
        "avg_tx_per_block": 0,
        "avg_value_per_tx_ETH": 0,
        "busiest_hour_utc": "",
    }
    return sentiment, news, chain, gov, evm


def test_build_context_structure_dedup_and_summary(monkeypatch):
    sentiment, news, chain, gov, evm = _dummy_components()
    snippets = ["previous proposal", "previous proposal"]
    monkeypatch.setattr(ollama_api, "generate_completion", lambda _p, **_: "summary")
    monkeypatch.setattr(proposal_store, "record_context", lambda _c: None)
    ctx = build_context(
        sentiment,
        news,
        chain,
        gov,
        evm,
        snippets,
        summarise_snippets=True,
    )
    assert set(ctx.keys()) == {
        "timestamp_utc",
        "sentiment",
        "news",
        "chain_kpis",
        "evm_kpis",
        "governance_kpis",
        "trending_topics",
        "kb_snippets",
        "kb_summary",
        "kb_embedded",
    }
    assert ctx["kb_snippets"] == ["previous proposal"]
    assert ctx["kb_summary"] == "summary"
    assert ctx["trending_topics"] == []
    assert v.validate_sentiment(ctx["sentiment"])
    assert v.validate_news(ctx["news"])
    assert v.validate_chain_kpis(ctx["chain_kpis"])
    assert v.validate_governance_kpis(ctx["governance_kpis"])
    assert v.validate_evm_kpis(ctx["evm_kpis"])


def test_search_proposals_by_keyword(monkeypatch):
    import pandas as pd

    df = pd.DataFrame(
        {"proposal_text": ["Increase staking rewards", "Governance overhaul", "Staking improvements"]}
    )
    monkeypatch.setattr(proposal_store, "load_proposals", lambda: df)
    res = proposal_store.search_proposals("staking", limit=5)
    assert res == ["Increase staking rewards", "Staking improvements"]


def test_record_context_persist(tmp_path, monkeypatch):
    sentiment, news, chain, gov, evm = _dummy_components()
    snippets = ["snippet"]
    monkeypatch.setattr(ollama_api, "generate_completion", lambda _p, **_: "summary")
    monkeypatch.setattr(proposal_store, "XLSX_PATH", tmp_path / "gov.xlsx")
    ctx = build_context(
        sentiment,
        news,
        chain,
        gov,
        evm,
        snippets,
        summarise_snippets=True,
    )
    from openpyxl import load_workbook

    wb = load_workbook(proposal_store.XLSX_PATH)
    ws = wb["Context"]
    rows = list(ws.iter_rows(values_only=True))
    # header then one row
    assert rows[0] == ("timestamp", "context_json")
    stored = json.loads(rows[1][1])
    assert stored["sentiment"] == sentiment
    assert stored["news"] == news
    assert stored["chain_kpis"] == chain
    assert stored["governance_kpis"] == gov
    assert stored["evm_kpis"] == evm
    assert stored["trending_topics"] == []
    assert stored["kb_snippets"] == snippets
    assert stored["kb_summary"] == "summary"


def test_component_weighting(monkeypatch):
    sentiment, news, chain, gov, evm = _dummy_components()
    sentiment["sentiment_score"] = 1
    chain["avg_tx_per_block"] = 10
    monkeypatch.setenv("DATA_WEIGHT_CHAT", "2")
    monkeypatch.setenv("DATA_WEIGHT_FORUM", "2")
    monkeypatch.setenv("DATA_WEIGHT_CHAIN", "0.5")
    evm["avg_tx_per_block"] = 10
    monkeypatch.setenv("DATA_WEIGHT_EVM", "0.5")
    monkeypatch.setattr(proposal_store, "record_context", lambda _c: None)
    ctx = build_context(sentiment, news, chain, gov, evm)
    assert ctx["sentiment"]["sentiment_score"] == 2
    assert ctx["chain_kpis"]["avg_tx_per_block"] == 5
    assert ctx["evm_kpis"]["avg_tx_per_block"] == 5
